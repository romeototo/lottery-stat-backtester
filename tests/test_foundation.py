from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys
import unittest
import zipfile
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from analyzer import data_quality_issues, data_quality_report, data_trust_score, load_lottery_file, prepare_dataframe, trust_warning_message
from backtester import compare_models_to_random_baseline, no_edge_warning_messages, random_baseline_simulation, random_baseline_summary, rolling_backtest, statistical_significance_table
from data_sources.glo_importer import load_glo_file, normalize_glo_dataframe
from exporter import build_excel_report


class NamedBytesIO(BytesIO):
    def __init__(self, data: bytes, name: str):
        super().__init__(data)
        self.name = name


class FoundationTests(unittest.TestCase):
    def test_streamlit_app_runs_without_exceptions(self):
        app = AppTest.from_file(str(ROOT / "app.py"))
        app.run(timeout=30)

        self.assertEqual(len(app.exception), 0)

    def test_sample_csv_loads_successfully(self):
        df = prepare_dataframe(pd.read_csv(ROOT / "data" / "sample_thai_lottery.csv", dtype=str))

        self.assertGreater(len(df), 0)
        self.assertIn("date", df.columns)
        self.assertIn("last_2_digits", df.columns)

    def test_load_lottery_file_accepts_csv_upload(self):
        payload = b"date,last_2_digits\n2024-01-17,09\n"
        uploaded = NamedBytesIO(payload, "sample.csv")

        df = load_lottery_file(uploaded)

        self.assertEqual(df.loc[0, "last_2_digits"], "09")

    def test_leading_zero_values_are_preserved(self):
        df = prepare_dataframe(
            pd.DataFrame(
                {
                    "date": ["2024-01-17", "2024-02-01", "2024-02-16"],
                    "last_2_digits": ["06", "09", "9"],
                }
            )
        )

        self.assertEqual(df["last_2_digits"].tolist(), ["06", "09", "09"])

    def test_data_quality_counts_invalid_values_correctly(self):
        df = prepare_dataframe(
            pd.DataFrame(
                {
                    "date": [
                        "2024-01-17",
                        "2024-01-17",
                        "not-a-date",
                        "2024-02-16",
                        "2024-03-01",
                    ],
                    "last_2_digits": ["06", "123", "09", "", "ab"],
                }
            )
        )

        report = data_quality_report(df)

        self.assertEqual(report["missing_date"], 1)
        self.assertEqual(report["missing_last_2_digits"], 1)
        self.assertEqual(report["duplicate_dates"], 1)
        self.assertEqual(report["invalid_last_2_digits"], 2)

    def test_data_quality_issue_report_lists_row_level_problems(self):
        df = prepare_dataframe(
            pd.DataFrame(
                {
                    "date": ["2024-01-17", "2024-01-17", "not-a-date", "2024-02-16"],
                    "last_2_digits": ["06", "123", "", "ab"],
                }
            )
        )

        issues = data_quality_issues(df)

        self.assertGreaterEqual(len(issues), 4)
        self.assertTrue(issues["issues"].str.contains("duplicate_date").any())
        self.assertTrue(issues["issues"].str.contains("missing_date").any())
        self.assertTrue(issues["issues"].str.contains("missing_last_2_digits").any())
        self.assertTrue(issues["issues"].str.contains("invalid_last_2_digits_format").any())

    def test_rolling_backtest_returns_results(self):
        df = prepare_dataframe(pd.read_csv(ROOT / "data" / "sample_thai_lottery.csv", dtype=str))

        details = rolling_backtest(df, "Hot Frequency", pick_count=5, warmup=5)

        self.assertGreater(len(details), 0)
        self.assertIn("hit", details.columns)

    def test_excel_export_creates_valid_file(self):
        df = prepare_dataframe(pd.read_csv(ROOT / "data" / "sample_thai_lottery.csv", dtype=str))

        excel_bytes = build_excel_report(df, pick_count=5, warmup=5, random_runs=100)

        self.assertGreater(len(excel_bytes), 0)
        self.assertTrue(zipfile.is_zipfile(BytesIO(excel_bytes)))

    def test_random_baseline_summary_has_required_statistics(self):
        baseline = random_baseline_simulation(draw_count=40, pick_count=5, runs=100)

        summary = random_baseline_summary(baseline)

        self.assertEqual(summary.loc[0, "runs"], 100)
        self.assertIn("mean_hit_rate_percent", summary.columns)
        self.assertIn("median_hit_rate_percent", summary.columns)
        self.assertIn("p05_hit_rate_percent", summary.columns)
        self.assertIn("p95_hit_rate_percent", summary.columns)

    def test_statistical_significance_output_has_p_value_and_ci(self):
        tournament = pd.DataFrame(
            {
                "model": ["Example"],
                "test_draws": [40],
                "hits": [4],
                "hit_rate_percent": [10.0],
            }
        )

        significance = statistical_significance_table(tournament, pick_count=5)

        self.assertIn("p_value_vs_random", significance.columns)
        self.assertIn("ci_low_percent", significance.columns)
        self.assertIn("ci_high_percent", significance.columns)
        self.assertIn("has_statistical_edge", significance.columns)

    def test_excel_export_includes_new_phase3_sheets(self):
        df = prepare_dataframe(pd.read_csv(ROOT / "data" / "sample_thai_lottery.csv", dtype=str))

        excel_bytes = build_excel_report(df, pick_count=5, warmup=5, random_runs=100)
        workbook = load_workbook(BytesIO(excel_bytes), read_only=True)

        self.assertIn("Data Quality Issues", workbook.sheetnames)
        self.assertIn("Random Baseline Summary", workbook.sheetnames)
        self.assertIn("Statistical Significance", workbook.sheetnames)

    def test_glo_style_import_mapping(self):
        raw = pd.DataFrame(
            {
                "วันที่ออกรางวัล": ["2024-01-17"],
                "รางวัลที่ 1": ["123456"],
                "รางวัลเลขท้าย 2 ตัว": ["09"],
                "เลขหน้า 3 ตัว 1": ["111"],
                "เลขหน้า 3 ตัว 2": ["222"],
                "เลขท้าย 3 ตัว 1": ["333"],
                "เลขท้าย 3 ตัว 2": ["444"],
            }
        )

        df = normalize_glo_dataframe(raw)

        self.assertEqual(df.loc[0, "last_2_digits"], "09")
        self.assertEqual(df.loc[0, "first_prize"], "123456")
        self.assertEqual(df.loc[0, "front_3_1"], "111")
        self.assertEqual(df.loc[0, "source"], "official_glo")

    def test_glo_thai_column_names_preserve_leading_zero_prizes(self):
        raw = pd.DataFrame(
            {
                "วันที่ออกรางวัล": ["17/01/2024"],
                "รางวัลที่ 1": ["012345"],
                "รางวัลเลขท้าย 2 ตัว": ["06"],
                "เลขหน้า 3 ตัว 1": ["009"],
                "เลขหน้า 3 ตัว 2": ["095"],
                "เลขท้าย 3 ตัว 1": ["007"],
                "เลขท้าย 3 ตัว 2": ["888"],
            }
        )

        df = normalize_glo_dataframe(raw)

        self.assertEqual(df.loc[0, "source"], "official_glo")
        self.assertEqual(df.loc[0, "first_prize"], "012345")
        self.assertEqual(df.loc[0, "last_2_digits"], "06")
        self.assertEqual(df.loc[0, "front_3_1"], "009")
        self.assertEqual(df.loc[0, "front_3_2"], "095")

    def test_glo_english_column_names_from_csv_upload(self):
        payload = (
            "draw_date,firstPrize,lastTwoDigits,front3First,front3Second,last3First,last3Second\n"
            "2024-01-17,123456,09,009,095,123,045\n"
        ).encode("utf-8")
        uploaded = NamedBytesIO(payload, "glo.csv")

        df = load_glo_file(uploaded)

        self.assertEqual(df.loc[0, "source"], "official_glo")
        self.assertEqual(df.loc[0, "last_2_digits"], "09")
        self.assertEqual(df.loc[0, "front_3_1"], "009")
        self.assertEqual(df.loc[0, "last_3_2"], "045")

    def test_glo_mixed_column_names_from_excel_upload(self):
        buffer = NamedBytesIO(b"", "glo.xlsx")
        raw = pd.DataFrame(
            {
                "period_date": ["2024-01-17"],
                "รางวัลที่ 1": ["654321"],
                "last_2": ["6"],
                "เลขหน้า 3 ตัว 1": ["9"],
                "front3Second": ["95"],
                "เลขท้าย 3 ตัว 1": ["7"],
                "last3Second": ["88"],
            }
        )
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            raw.to_excel(writer, index=False)
        buffer.seek(0)

        df = load_glo_file(buffer)

        self.assertEqual(df.loc[0, "last_2_digits"], "06")
        self.assertEqual(df.loc[0, "front_3_1"], "009")
        self.assertEqual(df.loc[0, "front_3_2"], "095")
        self.assertEqual(df.loc[0, "last_3_1"], "007")
        self.assertEqual(df.loc[0, "last_3_2"], "088")

    def test_glo_invalid_rows_and_duplicate_dates_reported(self):
        raw = pd.DataFrame(
            {
                "draw_date": ["2024-01-17", "2024-01-17", "bad-date"],
                "firstPrize": ["1234567", "123456", "abcdef"],
                "lastTwoDigits": ["123", "09", ""],
                "front3First": ["009", "095", "bad"],
            }
        )

        df = normalize_glo_dataframe(raw)
        quality = data_quality_report(df)
        issues = data_quality_issues(df)

        self.assertEqual(quality["duplicate_dates"], 1)
        self.assertEqual(quality["missing_date"], 1)
        self.assertEqual(quality["missing_last_2_digits"], 1)
        self.assertEqual(quality["invalid_last_2_digits"], 1)
        self.assertTrue(issues["issues"].str.contains("duplicate_date").any())

    def test_trust_score_high_medium_low(self):
        high = prepare_dataframe(
            pd.DataFrame(
                {
                    "date": pd.date_range("2020-01-01", periods=40, freq="15D").astype(str),
                    "last_2_digits": [f"{i % 100:02d}" for i in range(40)],
                    "source": ["GLO Open Data official"] * 40,
                }
            )
        )
        medium = prepare_dataframe(
            pd.DataFrame(
                {
                    "date": pd.date_range("2024-01-01", periods=20, freq="15D").astype(str),
                    "last_2_digits": [f"{i % 100:02d}" for i in range(20)],
                    "source": ["manual"] * 20,
                }
            )
        )
        low = prepare_dataframe(
            pd.DataFrame(
                {
                    "date": ["bad-date", "2024-01-17", "2024-01-17"],
                    "last_2_digits": ["", "123", "ab"],
                    "source": [""] * 3,
                }
            )
        )

        self.assertEqual(data_trust_score(high)[0]["trust_level"], "High trust")
        self.assertEqual(data_trust_score(medium)[0]["trust_level"], "Medium trust")
        self.assertEqual(data_trust_score(low)[0]["trust_level"], "Low trust")

    def test_low_trust_warning_logic(self):
        self.assertIn("Do not rely", trust_warning_message(40))
        self.assertEqual(trust_warning_message(85), "")

    def test_excel_export_includes_data_trust_score_sheet(self):
        df = prepare_dataframe(pd.read_csv(ROOT / "data" / "sample_thai_lottery.csv", dtype=str))

        excel_bytes = build_excel_report(df, pick_count=5, warmup=5, random_runs=100)
        workbook = load_workbook(BytesIO(excel_bytes), read_only=True)

        self.assertIn("Data Trust Score", workbook.sheetnames)

    def test_calibration_export_sheets_exist(self):
        df = prepare_dataframe(pd.read_csv(ROOT / "data" / "sample_thai_lottery.csv", dtype=str))

        excel_bytes = build_excel_report(df, pick_count=5, warmup=5, random_runs=100)
        workbook = load_workbook(BytesIO(excel_bytes), read_only=True)

        self.assertIn("Calibration Report", workbook.sheetnames)
        self.assertIn("Missing Draw Periods", workbook.sheetnames)
        self.assertIn("Invalid Prize Formats", workbook.sheetnames)

    def test_export_works_after_glo_style_import(self):
        raw = pd.DataFrame(
            {
                "draw_date": pd.date_range("2024-01-01", periods=12, freq="15D").astype(str),
                "firstPrize": [f"{i:06d}" for i in range(12)],
                "lastTwoDigits": [f"{i % 100:02d}" for i in range(12)],
                "front3First": [f"{i:03d}" for i in range(12)],
                "front3Second": [f"{(i + 1):03d}" for i in range(12)],
                "last3First": [f"{(i + 2):03d}" for i in range(12)],
                "last3Second": [f"{(i + 3):03d}" for i in range(12)],
            }
        )
        df = normalize_glo_dataframe(raw)

        excel_bytes = build_excel_report(df, pick_count=5, warmup=5, random_runs=100)
        workbook = load_workbook(BytesIO(excel_bytes), read_only=True)

        self.assertIn("Raw Data", workbook.sheetnames)
        self.assertIn("Data Trust Score", workbook.sheetnames)

    def test_no_edge_warning_appears_when_model_does_not_beat_baseline(self):
        tournament = pd.DataFrame(
            {
                "model": ["Weak Model"],
                "test_draws": [100],
                "hits": [3],
                "hit_rate_percent": [3.0],
            }
        )
        baseline_summary = pd.DataFrame([{"p95_hit_rate_percent": 8.0}])

        compared = compare_models_to_random_baseline(tournament, baseline_summary)
        messages = no_edge_warning_messages(compared)

        self.assertFalse(compared.loc[0, "beats_random_baseline"])
        self.assertTrue(any("No Edge" in message for message in messages))


if __name__ == "__main__":
    unittest.main()
